# pyright: reportOptionalSubscript=false, reportOptionalMemberAccess=false
"""Multi-format SOW export engine.
* docx  — hand-rolled OpenXML zip (Word & LibreOffice compatible)
* odt   — ODF text document (odfpy, pure Python)
* xlsx  — Excel workbook (openpyxl, pure Python)
* csv   — UTF-8 cost breakdown CSV
* xml   — MS Project-compatible WBS XML with PredecessorLink nodes
* md    — pass-through of stored Markdown
"""
from __future__ import annotations

import csv
import io
import re
import zipfile
from datetime import date, datetime, timedelta
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from app.models.schemas import SowResponse

COSTING_FORMATS = frozenset({"xlsx", "csv"})
ALL_FORMATS = ("md", "docx", "odt", "xlsx", "csv", "xml")
MIME_TYPES = {
    "md": "text/markdown; charset=utf-8",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "odt": "application/vnd.oasis.opendocument.text",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "xml": "application/xml; charset=utf-8",
}
DEFAULT_FILENAMES = {
    "md": "{title}.md",
    "docx": "{title}.docx",
    "odt": "{title}.odt",
    "xlsx": "{title}-cost-breakdown.xlsx",
    "csv": "{title}-cost-breakdown.csv",
    "xml": "{title}-gantt.xml",
}


def _coerce_sow(sow):
    if isinstance(sow, SowResponse):
        return sow
    return SowResponse.model_validate(sow)


def _safe_filename(title):
    return re.sub(r"[\\/:*?\"<>|]", "-", title or "SOW").strip() or "SOW"


def _esc(text):
    return ((text or "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


# ---------------------------------------------------------------------------
# .docx — hand-rolled OpenXML
# ---------------------------------------------------------------------------

def export_to_docx(sow, title):
    sow = _coerce_sow(sow)
    title_text = title or sow.project_title or "Scope of Work"
    body = _sow_to_docx_body(sow, title_text)
    return _pack_docx(body)


def _sow_to_docx_body(sow, title):
    def run(text):
        text = _esc(text)
        out = []
        for i, chunk in enumerate(re.split(r"\*\*(.+?)\*\*", text)):
            if i % 2 == 1:
                out.append("<w:r><w:rPr><w:b/></w:rPr><w:t>" + chunk + "</w:t></w:r>")
            elif chunk:
                out.append("<w:r><w:t>" + chunk + "</w:t></w:r>")
        return "".join(out)

    parts = ['<w:p><w:r><w:rPr><w:b/><w:sz w:val="32"/></w:rPr><w:t>' + _esc(title) + "</w:t></w:r></w:p>"]

    if sow.executive_summary:
        es = sow.executive_summary
        parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="28"/></w:rPr><w:t>Executive Summary</w:t></w:r></w:p>')
        if es.overview:
            parts.append("<w:p>" + run(es.overview) + "</w:p>")
        if es.overall_condition:
            parts.append('<w:p><w:r><w:rPr><w:b/></w:rPr><w:t xml:space="preserve">Overall condition: </w:t></w:r><w:r><w:t>' + _esc(es.overall_condition) + "</w:t></w:r></w:p>")
        if es.priority_findings:
            parts.append('<w:p><w:r><w:rPr><w:b/></w:rPr><w:t xml:space="preserve">Priority findings: </w:t></w:r><w:r><w:t>' + _esc(es.priority_findings) + "</w:t></w:r></w:p>")

    if sow.visual_findings:
        parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="28"/></w:rPr><w:t>Visual Findings</w:t></w:r></w:p>')
        rows = [["ID", "Asset", "Severity", "Description", "Action"]]
        for vf in sow.visual_findings:
            rows.append([vf.id or "", vf.asset or "", vf.severity or "", vf.description or "", vf.recommended_action or ""])
        parts.append(_docx_table(rows))

    if sow.recommended_services:
        parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="28"/></w:rPr><w:t>Recommended Services</w:t></w:r></w:p>')
        rows = [["ID", "Service", "Asset", "Priority", "Qty", "Unit", "Unit Cost", "Total"]]
        for s in sow.recommended_services:
            rows.append([s.id or "", s.service or "", s.asset or "", s.priority or "",
                         str(s.quantity), s.unit or "", f"{s.unit_cost:.2f}", f"{s.total_cost:.2f}"])
        parts.append(_docx_table(rows))

    if sow.scope_breakdown:
        parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="28"/></w:rPr><w:t>Scope of Work</w:t></w:r></w:p>')
        for scope in sow.scope_breakdown:
            parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="26"/></w:rPr><w:t>' + _esc(scope.phase or "Phase") + "</w:t></w:r></w:p>")
            if scope.work_description:
                parts.append("<w:p>" + run(scope.work_description) + "</w:p>")
            if scope.duration_days:
                parts.append('<w:p><w:r><w:rPr><w:i/></w:rPr><w:t xml:space="preserve">Duration: ' + str(int(scope.duration_days)) + " day(s)</w:t></w:r></w:p>")
            if scope.depends_on:
                parts.append('<w:p><w:r><w:rPr><w:i/></w:rPr><w:t xml:space="preserve">Depends on: ' + _esc(", ".join(scope.depends_on)) + "</w:t></w:r></w:p>")
            for d in (scope.deliverables or []):
                parts.append("<w:p><w:r><w:t>* " + _esc(d) + "</w:t></w:r></w:p>")

    if sow.cost_breakdown:
        cb = sow.cost_breakdown
        parts.append('<w:p><w:r><w:rPr><w:b/><w:sz w:val="28"/></w:rPr><w:t>Cost Breakdown</w:t></w:r></w:p>')
        rows = [["Item", f"Amount ({cb.currency or ''})"],
                ["Labor", f"{cb.labor:.2f}"], ["Materials", f"{cb.materials:.2f}"],
                ["Equipment", f"{cb.equipment:.2f}"], ["Subtotal", f"{cb.subtotal:.2f}"],
                [f"Contingency ({cb.contingency_pct}%)", f"{cb.contingency:.2f}"],
                ["Total", f"{cb.total:.2f}"]]
        parts.append(_docx_table(rows))

    parts.append("<w:p/>")
    return "".join(parts)


def _docx_table(rows):
    if not rows:
        return ""
    col_count = max(len(r) for r in rows)
    col_w = int(6200 / col_count)
    grid = "".join(f'<w:gridCol w:w="{col_w}"/>' for _ in range(col_count))
    out = ["<w:tbl><w:tblGrid>" + grid + "</w:tblGrid>"]
    for ri, row in enumerate(rows):
        out.append("<w:tr>")
        cells = list(row) + [""] * (col_count - len(row))
        for cell in cells:
            bold = "<w:rPr><w:b/></w:rPr>" if ri == 0 else ""
            out.append('<w:tc><w:tcPr><w:tcW w:w="' + str(col_w) + '" w:type="dxa"/></w:tcPr><w:p><w:r>' + bold + "<w:t>" + _esc(str(cell)) + "</w:t></w:r></w:p></w:tc>")
        out.append("</w:tr>")
    out.append("</w:tbl><w:p/>")
    return "".join(out)


def _pack_docx(body_xml):
    doc_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>" + body_xml +
        '<w:sectPr><w:pgSz w:w="12240" w:h="15840"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/></w:sectPr>'
        "</w:body></w:document>"
    )
    ct = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '<Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="word/styles.xml"/>'
        "</Relationships>"
    )
    doc_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        "</Relationships>"
    )
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"/>'
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ct)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/_rels/document.xml.rels", doc_rels)
        z.writestr("word/document.xml", doc_xml)
        z.writestr("word/styles.xml", styles_xml)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# .odt — ODF text document via odfpy
# ---------------------------------------------------------------------------

def export_to_odt(sow, title):
    from odf.opendocument import OpenDocumentText
    from odf.text import H, P

    sow = _coerce_sow(sow)
    title = title or sow.project_title or "Scope of Work"
    doc = OpenDocumentText()
    doc.text.addElement(H(outlinelevel=1, text=title))

    if sow.executive_summary:
        es = sow.executive_summary
        doc.text.addElement(H(outlinelevel=2, text="Executive Summary"))
        if es.overview:
            doc.text.addElement(P(text=es.overview))
        if es.overall_condition:
            doc.text.addElement(P(text=f"Overall condition: {es.overall_condition}"))
        if es.priority_findings:
            doc.text.addElement(P(text=f"Priority findings: {es.priority_findings}"))

    if sow.visual_findings:
        doc.text.addElement(H(outlinelevel=2, text="Visual Findings"))
        for vf in sow.visual_findings:
            doc.text.addElement(P(text=f"[{vf.id}] {vf.asset} -- {vf.severity}: {vf.description}"))
            if vf.recommended_action:
                doc.text.addElement(P(text=f"   Action: {vf.recommended_action}"))

    if sow.recommended_services:
        doc.text.addElement(H(outlinelevel=2, text="Recommended Services"))
        for s in sow.recommended_services:
            line = (f"[{s.id}] {s.service} ({s.asset}) -- {s.priority}, "
                    f"qty {s.quantity} {s.unit} @ {s.unit_cost:.2f} = {s.total_cost:.2f}")
            doc.text.addElement(P(text=line))
            if s.notes:
                doc.text.addElement(P(text=f"   Notes: {s.notes}"))

    if sow.scope_breakdown:
        doc.text.addElement(H(outlinelevel=2, text="Scope of Work"))
        for scope in sow.scope_breakdown:
            doc.text.addElement(H(outlinelevel=3, text=scope.phase or "Phase"))
            if scope.work_description:
                doc.text.addElement(P(text=scope.work_description))
            if scope.duration_days:
                doc.text.addElement(P(text=f"Duration: {int(scope.duration_days)} day(s)"))
            if scope.depends_on:
                doc.text.addElement(P(text=f"Depends on: {', '.join(scope.depends_on)}"))
            for d in (scope.deliverables or []):
                doc.text.addElement(P(text=f"  * {d}"))

    if sow.cost_breakdown:
        cb = sow.cost_breakdown
        doc.text.addElement(H(outlinelevel=2, text="Cost Breakdown"))
        currency = cb.currency or ""
        for line in [
            f"Labor:      {cb.labor:.2f} {currency}",
            f"Materials:  {cb.materials:.2f} {currency}",
            f"Equipment:  {cb.equipment:.2f} {currency}",
            f"Subtotal:   {cb.subtotal:.2f} {currency}",
            f"Contingency ({cb.contingency_pct}%): {cb.contingency:.2f} {currency}",
            f"Total:      {cb.total:.2f} {currency}",
        ]:
            doc.text.addElement(P(text=line))

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# .xlsx — Excel workbook via openpyxl
# ---------------------------------------------------------------------------

def export_to_xlsx(sow, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sow = _coerce_sow(sow)
    title_text = title or sow.project_title or "SOW"
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    ws1 = wb.active
    ws1.title = "Cost Summary"
    cb = sow.cost_breakdown
    currency = cb.currency or ""
    rows1 = [
        ("Project", title_text),
        ("Site", sow.site or ""),
        ("Client", sow.client or ""),
        ("Generated at", sow.generated_at or ""),
        ("", ""),
        ("Item", f"Amount ({currency})"),
        ("Labor", round(cb.labor, 2)),
        ("Materials", round(cb.materials, 2)),
        ("Equipment", round(cb.equipment, 2)),
        ("Subtotal", round(cb.subtotal, 2)),
        ("Contingency %", round(cb.contingency_pct, 2)),
        ("Contingency", round(cb.contingency, 2)),
        ("Total", round(cb.total, 2)),
    ]
    for r, (a, b) in enumerate(rows1, start=1):
        ws1.cell(row=r, column=1, value=a)
        ws1.cell(row=r, column=2, value=b)
        if r == 6:
            ws1.cell(row=r, column=1).font = header_font
            ws1.cell(row=r, column=1).fill = header_fill
            ws1.cell(row=r, column=2).font = header_font
            ws1.cell(row=r, column=2).fill = header_fill
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    ws2 = wb.create_sheet("Services")
    headers2 = ["ID", "Service", "Asset", "Priority", "Qty", "Unit", "Unit Cost", "Total Cost", "Notes"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r, s in enumerate(sow.recommended_services, start=2):
        ws2.cell(row=r, column=1, value=s.id)
        ws2.cell(row=r, column=2, value=s.service)
        ws2.cell(row=r, column=3, value=s.asset)
        ws2.cell(row=r, column=4, value=s.priority)
        ws2.cell(row=r, column=5, value=s.quantity)
        ws2.cell(row=r, column=6, value=s.unit)
        ws2.cell(row=r, column=7, value=round(s.unit_cost, 2))
        ws2.cell(row=r, column=8, value=round(s.total_cost, 2))
        ws2.cell(row=r, column=9, value=s.notes or "")
    for col_letter, width in zip("ABCDEFGHI", [8, 32, 16, 10, 8, 8, 12, 12, 32]):
        ws2.column_dimensions[col_letter].width = width

    ws3 = wb.create_sheet("Scope Breakdown")
    headers3 = ["Seq", "Phase", "Work Description", "Duration (days)", "Depends On", "Deliverables"]
    for col, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r, scope in enumerate(sow.scope_breakdown, start=2):
        ws3.cell(row=r, column=1, value=scope.sequence or (r - 1))
        ws3.cell(row=r, column=2, value=scope.phase)
        ws3.cell(row=r, column=3, value=scope.work_description)
        ws3.cell(row=r, column=4, value=scope.duration_days or 0)
        ws3.cell(row=r, column=5, value=", ".join(scope.depends_on or []))
        ws3.cell(row=r, column=6, value="\n".join(scope.deliverables or []))
    for col_letter, width in zip("ABCDEF", [6, 30, 50, 14, 28, 40]):
        ws3.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# .csv — UTF-8 cost breakdown
# ---------------------------------------------------------------------------

def export_to_csv(sow, title):
    sow = _coerce_sow(sow)
    title_text = title or sow.project_title or "SOW"
    cb = sow.cost_breakdown
    currency = cb.currency or ""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["OSIRIS Imhotep -- SOW Cost Breakdown"])
    writer.writerow(["Project", title_text])
    writer.writerow(["Site", sow.site or ""])
    writer.writerow(["Client", sow.client or ""])
    writer.writerow(["Generated at", sow.generated_at or ""])
    writer.writerow([])
    writer.writerow(["Item", f"Amount ({currency})"])
    writer.writerow(["Labor", f"{cb.labor:.2f}"])
    writer.writerow(["Materials", f"{cb.materials:.2f}"])
    writer.writerow(["Equipment", f"{cb.equipment:.2f}"])
    writer.writerow(["Subtotal", f"{cb.subtotal:.2f}"])
    writer.writerow([f"Contingency ({cb.contingency_pct}%)", f"{cb.contingency:.2f}"])
    writer.writerow(["Total", f"{cb.total:.2f}"])
    writer.writerow([])
    writer.writerow(["ID", "Service", "Asset", "Priority", "Qty", "Unit", "Unit Cost", "Total Cost", "Notes"])
    for s in sow.recommended_services:
        writer.writerow([
            s.id or "", s.service or "", s.asset or "", s.priority or "",
            s.quantity, s.unit or "", f"{s.unit_cost:.2f}", f"{s.total_cost:.2f}", s.notes or ""
        ])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# ---------------------------------------------------------------------------
# .xml — MS Project-compatible WBS / Gantt
# ---------------------------------------------------------------------------

def export_to_gantt_xml(sow, title):
    sow = _coerce_sow(sow)
    title_text = title or sow.project_title or "SOW"
    start_date = _parse_iso_date(sow.generated_at) or date.today()
    phases = list(sow.scope_breakdown or [])
    phases.sort(key=lambda p: (p.sequence or 0, 0))
    phase_uid = {}
    for idx, p in enumerate(phases, start=1):
        phase_uid[p.phase or f"Phase {idx}"] = idx
    NS = "http://schemas.microsoft.com/project"
    ET.register_namespace("", NS)
    project = ET.Element(f"{{{NS}}}Project")
    ET.SubElement(project, f"{{{NS}}}Name").text = title_text
    sv = ET.SubElement(project, f"{{{NS}}}SaveVersion")
    sv.text = "14"
    tasks_container = ET.SubElement(project, f"{{{NS}}}Tasks")
    for idx, p in enumerate(phases, start=1):
        name = p.phase or f"Phase {idx}"
        prior = phases[: idx - 1]
        offset = sum(int(x.duration_days or 0) for x in prior)
        task_start = start_date + timedelta(days=offset)
        duration = int(p.duration_days or 0)
        task_finish = task_start + timedelta(days=max(duration - 1, 0)) if duration > 0 else task_start
        task = ET.SubElement(tasks_container, f"{{{NS}}}Task")
        ET.SubElement(task, f"{{{NS}}}UID").text = str(idx)
        ET.SubElement(task, f"{{{NS}}}ID").text = str(idx)
        ET.SubElement(task, f"{{{NS}}}Name").text = name
        ET.SubElement(task, f"{{{NS}}}Type").text = "0"
        ET.SubElement(task, f"{{{NS}}}IsManual").text = "0"
        dur_h = duration * 8
        ET.SubElement(task, f"{{{NS}}}Duration").text = f"PT{dur_h}H0M0S"
        ET.SubElement(task, f"{{{NS}}}DurationFormat").text = "7"
        ET.SubElement(task, f"{{{NS}}}Start").text = task_start.strftime("%Y-%m-%dT00:00:00")
        ET.SubElement(task, f"{{{NS}}}Finish").text = task_finish.strftime("%Y-%m-%dT00:00:00")
        ET.SubElement(task, f"{{{NS}}}Work").text = f"PT{dur_h}H0M0S"
        ET.SubElement(task, f"{{{NS}}}Sequence").text = str(p.sequence or idx)
        if p.work_description:
            ET.SubElement(task, f"{{{NS}}}Notes").text = p.work_description
        deps = [phase_uid[d] for d in (p.depends_on or []) if d in phase_uid]
        if deps:
            links = ET.SubElement(task, f"{{{NS}}}PredecessorLink")
            for d_uid in deps:
                link = ET.SubElement(links, f"{{{NS}}}Predecessor")
                ET.SubElement(link, f"{{{NS}}}UID").text = str(d_uid)
                ET.SubElement(link, f"{{{NS}}}PredecessorUID").text = str(d_uid)
                ET.SubElement(link, f"{{{NS}}}Type").text = "1"
                ET.SubElement(link, f"{{{NS}}}CrossProject").text = "0"
                ET.SubElement(link, f"{{{NS}}}LinkLag").text = "0"
                ET.SubElement(link, f"{{{NS}}}LagFormat").text = "7"
    ET.indent(project, space="  ")
    return ET.tostring(project, encoding="utf-8", xml_declaration=True)


def _parse_iso_date(value):
    if not value:
        return None
    s = value.strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S.%fZ"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Markdown — pass-through
# ---------------------------------------------------------------------------

def export_to_markdown(content_md):
    return (content_md or "").encode("utf-8")


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------

def export_sow(sow, content_md, title, formats):
    sow = _coerce_sow(sow)
    safe = _safe_filename(title or sow.project_title or "SOW")
    out = {}
    for fmt in formats:
        fmt = fmt.lower().strip()
        if fmt == "md":
            out[fmt] = (DEFAULT_FILENAMES["md"].format(title=safe), export_to_markdown(content_md))
        elif fmt == "docx":
            out[fmt] = (DEFAULT_FILENAMES["docx"].format(title=safe), export_to_docx(sow, title))
        elif fmt == "odt":
            out[fmt] = (DEFAULT_FILENAMES["odt"].format(title=safe), export_to_odt(sow, title))
        elif fmt == "xlsx":
            out[fmt] = (DEFAULT_FILENAMES["xlsx"].format(title=safe), export_to_xlsx(sow, title))
        elif fmt == "csv":
            out[fmt] = (DEFAULT_FILENAMES["csv"].format(title=safe), export_to_csv(sow, title))
        elif fmt == "xml":
            out[fmt] = (DEFAULT_FILENAMES["xml"].format(title=safe), export_to_gantt_xml(sow, title))
        else:
            raise ValueError(f"Unsupported export format: {fmt!r}")
    return out
